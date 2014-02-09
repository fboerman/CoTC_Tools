__author__ = 'williewonka'

import openpyxl
import json
import os
import encoding

existing_provinces = []
#prepare a list of exisisting province files
for info in os.walk('provinces'):
    files = info[2]
    break
for filename in files:
    if len(filename.split("-")) == 2:
        provincename = filename.split("-")[1].split(".")[0].strip()
    else:
        items = filename.split("-")
        items.pop(0)
        provincename = "-".join(items).split(".")[0].strip()
    existing_provinces.append(provincename)

#load the json file for hierarchy
jsonfile = open("provinces_hierarchy.json", "r")
hierarchy = json.loads(jsonfile.readlines()[0])

#load the extisting provinces info json
jsonfile = open("provinces.json", "r")
existing_provinces_info = json.loads(jsonfile.readlines()[0])

#create the excel workbook
wb = openpyxl.Workbook()
sheet = wb.get_active_sheet()
#write the headers
sheet.cell(row=0, column=0).value = "kingdom"
sheet.cell(row=0, column=1).value = "duchy"
sheet.cell(row=0, column=2).value = "county"
sheet.cell(row=0, column=3).value = "culture"
sheet.cell(row=0, column=4).value = "ideology"
sheet.cell(row=0, column=5).value = "has file already"
sheet.cell(row=0, column=6).value = "number of total holding slots"
sheet.cell(row=0, column=7).value = "barony"
sheet.cell(row=0, column=8).value = "holding type"
sheet.cell(row=0, column=9).value = "spacestation type (0-4)"
sheet.cell(row=0, column=10).value = "Atmosphere type (0-6)"
sheet.cell(row=0, column=11).value = "Water type (0-4)"
sheet.cell(row=0, column=12).value = "Temperature type (0-7)"
sheet.cell(row=0, column=13).value = "Colony type (0-4)"
sheet.cell(row=0, column=14).value = "Astroids type (blank or 1)"

#iterate trough the hierarchy and print all of the info
row = 1
for kingdomid in hierarchy:
    kingdom = hierarchy[kingdomid]
    sheet.cell(row=row,column=0).value = kingdom['name']
    row += 1
    for duchyid in kingdom['children']:
        duchy = kingdom['children'][duchyid]
        sheet.cell(row=row,column=1).value = duchy['name']
        row += 1
        for countyid in duchy['children']:
            #get the county name and print it
            county = duchy['children'][countyid]
            sheet.cell(row=row,column=2).value = county['name']
            #if there is already a file get the info from that
            if county['name'] in existing_provinces:
                sheet.cell(row=row, column=5).value = 1
                info = existing_provinces_info[county['name']]
                sheet.cell(row=row,column=3).value = info['culture']
                sheet.cell(row=row,column=4).value = info['religion']
                sheet.cell(row=row,column=6).value = info['max_settlements']
                #iterate through the baronies if they exist
                if 'baronies' in info:
                    for baronie_name in info['baronies'].keys():
                        #basic barony info
                        baronie_info = info['baronies'][baronie_name]
                        row += 1
                        sheet.cell(row=row, column=7).value = baronie_name
                        sheet.cell(row=row,column=8).value = baronie_info['holding_type']
                        #check all the building levels
                        if 'Atmosphere' in baronie_info:
                            sheet.cell(row=row, column=10).value = baronie_info['Atmosphere']
                        if 'Temperature' in baronie_info:
                            sheet.cell(row=row, column=12).value = baronie_info['Temperature']
                        if 'Water' in baronie_info:
                            sheet.cell(row=row, column=11).value = baronie_info['Water']
                        if 'SpaceStation' in baronie_info:
                            sheet.cell(row=row, column=9).value = baronie_info['SpaceStation']
                        if 'Colony' in baronie_info:
                            sheet.cell(row=row, column=13).value = baronie_info['Colony']
                        if 'Asteroids' in baronie_info:
                            sheet.cell(row=row, column=14).value = baronie_info['Asteroids']
            else:
                sheet.cell(row=row, column=5).value = 0
                #no existing file so get religion and culture from defined areas
                #first find culture
                if kingdom['name'] in encoding.cultures_per_region:
                    culture = encoding.cultures_per_region[kingdom['name']]
                elif duchy['name'] in encoding.cultures_per_region:
                    culture = encoding.cultures_per_region[duchy['name']]
                else:
                    #no culture defined
                    culture = 'Undefined'
                sheet.cell(row=row,column=3).value = culture
                #now find religion
                if kingdom['name'] in encoding.ideology_per_region:
                    religion = encoding.ideology_per_region[kingdom['name']]
                elif duchy['name'] in encoding.ideology_per_region:
                    religion = encoding.ideology_per_region[duchy['name']]
                else:
                    religion = 'Undefined'
                sheet.cell(row=row,column=4).value = religion
            row += 1

wb.save("Provinces_Attributes.xlsx")
print("done")