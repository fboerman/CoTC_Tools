__author__ = 'williewonka'
__excelfile__ = 'Provinces_Attributes.xlsx'

import openpyxl
import json
import os
import encoding
import argparse

existing_provinces = {}
maxid = 0
# prepare a dictonionary with existing province names and id numbers
for info in os.walk('provinces'):
    files = info[2]
    break
for filename in files:
    if len(filename.split("-")) == 2:
        provincename = filename.split("-")[1].split(".")[0].strip()
        province_id = filename.split("-")[0].strip()
    else:
        items = filename.split("-")
        province_id = items.pop(0)
        provincename = "-".join(items).split(".")[0].strip()
    existing_provinces[provincename] = province_id
    if int(province_id) > maxid:
        maxid = int(province_id)

def Create_Json():
    #load the excel file and sheet
    wb = openpyxl.load_workbook(__excelfile__)
    sheet = wb.get_active_sheet()

    #iterate trhough the file, skip the first line with headers

    row = 1

    hierarchy = {}

    provinces = {}


    while row < len(sheet.rows):
        #check if row is kingdom entry
        if sheet.cell(row=row,column=0).value != "" and sheet.cell(row=row,column=0).value != 'x':
            #create the kingdomentry and iterate through it
            kingdomname = sheet.cell(row=row, column=0).value
            hierarchy[kingdomname] = {}
            row += 1
            print('row: ' + str(row))
            while row < len(sheet.rows):#iterate through the kingdom
                if sheet.cell(row=row,column=0).value is not None and sheet.cell(row=row,column=0).value != 'x':#check if not a kingdom is present, which ends current one
                    row -= 1
                    break
                #check for duchy
                if sheet.cell(row=row,column=1).value is not None and sheet.cell(row=row,column=1).value != 'x':
                    #create duchy entry and iterate trought it
                    duchyname = sheet.cell(row=row,column=1).value
                    hierarchy[kingdomname][duchyname] = []
                    row += 1
                    print('row: ' + str(row))
                    while row < len(sheet.rows):#iterate throught the county list
                        if (sheet.cell(row=row,column=1).value is not None and sheet.cell(row=row,column=1).value != 'x') or (sheet.cell(row=row, column=0).value is not None and sheet.cell(row=row, column=0).value != 'x'): #check if a duchy is present, which ends the current one
                            row -= 1
                            break
                        #check for a county and save the info
                        if sheet.cell(row=row, column=2).value is not None and sheet.cell(row=row, column=2).value != 'x':
                            #save the name into the hierarchy
                            hierarchy[kingdomname][duchyname].append(sheet.cell(row=row, column=2).value)
                            #build the info except baronie into a province data object
                            name = sheet.cell(row=row, column=2).value
                            province = {
                                'title' : 'c_' + name.lower().replace(' ', '_'),
                                'culture' : sheet.cell(row=row, column=3).value,
                                'religion' : sheet.cell(row=row, column=4).value,
                                'baronies' : {}
                            }
                            if sheet.cell(row=row, column=6).value is not None:
                                province['max_settlements'] = sheet.cell(row=row, column=6).value
                            #iterate trough the baronies and put it in the province data object
                            row += 1
                            print('row: ' + str(row))
                            while row < len(sheet.rows):
                                if sheet.cell(row=row, column=7).value is None and sheet.cell(row=row, column=7).value != 'x':
                                    row -= 1
                                    break
                                baronyname = sheet.cell(row=row, column=7).value
                                province['baronies'][baronyname] = {
                                    'holding_type' : sheet.cell(row=row, column=8).value
                                }
                                def check_cell(column):#checks if given cell has valid entry, returns cell otherwise None
                                    if sheet.cell(row=row, column=column).value is not None or sheet.cell(row=row, column=column).value != 'x':
                                        return sheet.cell(row=row, column=column).value
                                    else:
                                        return None
                                def check_building(column, name):
                                    value = check_cell(column)
                                    if value is not None:
                                        province['baronies'][baronyname][name] = value

                                #check all the buildings
                                check_building(9,'SpaceStation')
                                check_building(10,'Athmosphere')
                                check_building(11,'Water')
                                check_building(12,'Temperature')
                                check_building(13,'Colony')
                                check_building(14,'Asteroids')

                                row += 1
                                print('row: ' + str(row))
                            #save the province dataobject
                            provinces[name] = province
                        row += 1
                        print('row: ' + str(row))

                row += 1
                print('row: ' + str(row))


        row += 1
        print('row: ' + str(row))

    #save to json object
    provinces_json = json.dumps(provinces)
    hierarchy_json = json.dumps(hierarchy)
    #first create the export directory
    if not os.path.exists('./export'):
        os.makedirs('./export')

    #save to file
    stream = open('./export/provinces.json', 'w')
    stream.writelines(provinces_json)
    stream.close()
    stream = open('./export/hierarchy.json', 'w')
    stream.writelines(hierarchy_json)
    stream.close()
    print("wrote json to file")
    return provinces

def Create_GameFiles(provinces):
    global maxid
    #first create the export directory
    if not os.path.exists('./export/provinces'):
        os.makedirs('./export/provinces')
    print("creating game files")
    #iterate through the json and create the province files
    count = 0
    for countyname in list(provinces.keys()):
        count += 1
        countyinfo = provinces[countyname]
        #check if province has existing id
        if countyname in existing_provinces:
            p_id = existing_provinces[countyname]
        else:
            maxid += 1
            p_id = maxid
        print('province ' + str(count) + ' of ' + str(len(provinces)))
        filename = str(p_id).strip() + ' - ' + countyname + '.txt'
        #open the file and start building it
        stream = open('./export/provinces/' + filename, 'w')
        stream.write('# ' + filename + '\n\n')
        stream.write('# County Title\n')
        stream.write('title = ' + countyinfo['title'] + '\n\n')
        stream.write('# Settlements\n')
        try:
            stream.write('max_settlements = ' + str(countyinfo['max_settlements']) + '\n')
        except KeyError:#if non existing just skip the line
            pass
        for baronyname in countyinfo['baronies']:
            stream.write(baronyname + ' = ' + countyinfo['baronies'][baronyname]['holding_type'] + '\n')
        stream.write('\n# Misc\n')
        if countyinfo['culture'] != 'Undefined':
            stream.write('culture = ' + countyinfo['culture'] + ' # Initial culture\n')
        if countyinfo['religion'] != 'Undefined':
            stream.write('religion = ' + countyinfo['religion'] + '\n')
        stream.write('\n# History\n')
        stream.write('1.1.1 = {\n')
        for baronyname in countyinfo['baronies']:
            for building in list(countyinfo['baronies'][baronyname].keys()):
                if building == 'holding_type':
                    continue
                encoding_options = {
                    'Athmosphere' : encoding.Atmosphere,
                    'Water' : encoding.Water,
                    'Temperature' : encoding.Temperature,
                    'SpaceStation' : encoding.SpaceStation,
                    'Asteroids' : encoding.Asteroids,
                    'Colony' : encoding.Colony
                }
                buildingcode = encoding_options[building][countyinfo['baronies'][baronyname][building]].replace('xx',
                    encoding.holding_codes[countyinfo['baronies'][baronyname]['holding_type']])
                stream.write('\t' + baronyname + ' = ' + buildingcode + '\n')
        stream.write('}\n\n')
        #if the file already existed, copy the history after 1.1.1 from existing file
        if countyname in existing_provinces:
            write = False
            for line in open('./provinces/' + filename,'r',encoding="UTF-8"):
                if line.strip() == '2998.1.1 = {':
                    write = True
                    stream.write(line)
                elif write:
                    stream.write(line)
        #done writing file
        stream.close()

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description="tool that imports data from excel file and converts them to a json object and gamefiles")
    parser.add_argument('--operation', nargs='?', const=1, type=str, default='both', help='json/gamefiles/both')

    choice = parser.parse_args().operation
    if choice == 'both':
        Create_GameFiles(Create_Json())
    elif choice == 'gamefiles':
        jsonfile = open("./export/provinces.json", "r")
        Create_GameFiles(json.loads(jsonfile.readlines()[0]))
    elif choice == 'json':
        Create_Json()
    else:
        print("Invalid choice!")